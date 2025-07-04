"""
Version 1.8

transfers drop off and pick up times from one Excel file to another that then calculates
the appropriate amount of money to charge.

General features
 1 -- This version transfers time from one Excel file to another.
 2 -- There is a window interface to choose the files needed in this script.
 3 -- If there are missing files from the reference data (data downloaded from hugh note) it will notify you that not
      all files have been downloaded from hugnote, but will transfer data with what ever files are available.
 4 -- If children on the hugnote files(reference files) cannot be found in the 預かり料金表 then it will notify you which
      children cannot be found along with the class name.
 5 -- Converts the pickup time if a child is in 課外授業 and is 一号.  Children that are 一号 taking the after school class
      are exempt for charges resulting for being picked up late, up to a certain point.
 6 -- Created function to replace spaces between names including the japanese space aka IDEOGRAPHIC SPACE character or
      \u3000 in unicode escape character.  This was done to reduce replication to reduce effort when refactoring code.
 7 -- Iterate through the Excel file to find where we charged extra money and fill in those cells with a pink color,
      to make it easier to find where we charged extra.
 8 -- Fixed it so that the workbooks are properly closed at the end of the function to prevent any unwanted things
      from happening with other functions down the line.
 9 -- Fixed the issue where the VBA code needed to be recalculated by opening the Excel file in Excel by triggering the
      recalculation within python.  Also made it so that the Excel opening up is invisible to make it cleaner.
10 -- No longer need to physically choose the recalculated Excel file during execution, it is automatically passed into
      the function that fills in the cells with extra charges.
11 -- Made it so the reference files (Excel documents downloaded from hug note that has the time stamps of arrival and
      departure times of all the kids) can be opened regardless if they are zipped or unzipped. (currently the default
      behavior is to choose a zipped file)
12 -- Fixed the issue where dep_check_time was being applied to all children. We only want to apply this to
      children that are 一号.
13 -- Iterate through the Excel file to fill in cells that have both the arrival time and departure time blank with
      休み to indicate that the child did not come to school on that day. Also highlight with yellow on cells that have
      only arrival time or departure time missing but not both to indicate something went wrong or the parents forgot
      to record the time for arrival or departure.
(new)
14 --
(working on)
** -- Finish type hints and doc strings for all the functions.
** -- Fixed issue where 一号課外 time adjustments were being made every single week. It's not every week that they have
      課外 classes, some are twice a month and some get canceled for one reason or another.
** -- Cleaned up code so that 0 index and 1 index difference between enumerate and the workbook are taken care of within
      their respective functions.
** -- Other various cleanups to make the code readable as well as organize things and changes to speed up things.
"""
from __future__ import annotations
from typing import Union
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from datetime import datetime
import openpyxl
from openpyxl import Workbook
# from openpyxl import styles
from openpyxl.styles import PatternFill
import xlwings as xw
import zipfile
from io import StringIO
import os
from openpyxl.worksheet.worksheet import Worksheet


def replace_all_spaces(words: str) -> str:
    # noinspection GrazieInspection
    """
        Remove blank space, whether it is english space or the japanese space.

        :param words: a string that may or may not contain a blank space.
        :return: Return a string with english and japanese spaces removed.
                 (may still contain any other form of blank spaces)
        """
    words = words.replace('\u3000', '')  # \u3000 is the equivalent to the japanese space. normal space -> ' '
    words = words.replace(' ', '')  # japanese space -> '　'
    return words


def find_date(tab: Worksheet, date: datetime) -> Union[list[int], None]:
    # noinspection GrazieInspection
    """
        Find the row and column (essentially the coordinates) of the matching date.

        :param tab: The current worksheet in use
        :param date: Datetime object from the reference file
        :return: A list of two integers representing the row and column the datetime had a match.
                 If no match was found return None.
        """
    for i, row in enumerate(tab.iter_rows()):
        for idx, cell in enumerate(row):
            if cell.value == date:
                return [i, idx]
    return None


# FIXME: Fix the function so that it doesnt have to return a list object and just an integer just for clarity sake. this
# FIXME: will involve fixing other functions that uses this out put as its input. (check later for the names
#        of those functions)
# specifically returning one int in the format of a list to avoid out of index errors.
def find_name(tab: Worksheet, name: str, date_row: int) -> list[int]:
    # noinspection GrazieInspection
    """
        Find the row number of where the child's name is located in the workbook.  Generally speaking we expect to find
        two locations, but sometimes we don't find them at all due to the kanji being the incorrect one and a mismatch
        happening from the reference file downloaded from hugnote and the record keeping file in Excel.
        ex: 髙田　!= 高田　though they seem similar they are two completely different strings in unicode.

        :param tab: the current Worksheet we are iterating through
        :param name: the name of the child we are looking for in the Worksheet.
        :param date_row: the first number in the list that is returned from the find_date function representing the row
                         in which the date was found.
        :return:
        """
    ans = []
    for i, row in enumerate(tab.iter_rows()):
        if isinstance(row[2].value, str):
            cell_name = replace_all_spaces(row[2].value)
            if cell_name == name:
                ans.append(i)
    if date_row < 10:
        return ans[:1]  # returning just one to make sure we only place the time correctly for the corresponding date
    else:
        return ans[-1:]  # also using [:1] and [-1:] so an error is not raised when the list is empty


def arr_check_time(time: int) -> int:
    """Convert the arrival time so that if anyone arrives before 7:30 it is set to 7:30"""
    if time < 730:
        time = 730
    return time


def dep_check_time(time: int) -> int:
    """Convert the departure time so that if they are 9 minutes over the time limit it reverts to the time limit
        So that they are not charged"""
    if 1131 <= time <= 1139:
        time = 1130
    if 1431 <= time <= 1439:
        time = 1430
    return time


# todo added flag so that if there is an empty number instead of 1, 2 or 3 it notifies the user instead of crashing.
# todo not i need to implement a dialog that opens up when such a case is recognized. instead of just a print out in the
# todo terminal.
def ichigo_check(name_coor: list[int], sheet: Worksheet) -> bool:
    # noinspection GrazieInspection
    """
        returning a bool to check if a given child is in the 一号 category. There is a cell in the workbook that denotes
        this information
        :param name_coor: A list containing a single integer which tells you which row we need to look at. we will be
                          looking at the first cell of the row.
        :param sheet: The worksheet we are currently iterating through.
        :return:  A bool value indicating whether the child is in the 一号 category.
        """
    row = name_coor[0] + 1  # adjust
    col = 1
    value = sheet.cell(row=row, column=col).value
    if isinstance(value, int):
        value = int(sheet.cell(row=row, column=col).value)
    else:
        print(f'Value is not an integer, is it a {type(value)} type and its value is {value}')
        #print(name_coor)
        print(f'{sheet.cell(row=name_coor[0]+1, column=3).value}が何号が入力されてない可能性があります。')

    if value == 1:
        return True
    else:
        return False


# TODO use iter_rows to iterate though the sheet instead of accessing it directly with slicing
def kagai_ichigo_check_time(name: str, time: int, day_of_week: int, sheet: Worksheet, date: datetime) -> int:
    """
    Convert departure time for children that are in 課外授業 and are 一号 to 1430 if they have class that day, and
    they leave before the pickup time limit.
    """
    days_of_week = ["月曜日", "火曜日", "水曜日", "木曜日", "金曜日", "土曜日", "日曜日"]
    for i, row in enumerate(sheet.iter_rows()):
        name_val = row[1].value
        if name_val is not None:
            name_val = replace_all_spaces(name_val)
        if name == name_val:
            time_limit = sheet[i+1][2 + day_of_week].value
            if time_limit is None:
                # print('修正なし: {}は{}に課外がありません。'.format(name, days_of_week[day_of_week]))
                return time
            elif 1430 < time <= time_limit:
                print('修正あり:　{} {} --> 1430。'.format(name, time))
                time = 1430
                return time
            elif time > time_limit:
                print('修正なし: {}{}が限度の{}を超えてるため。'.format(name, time, time_limit))
                return time
    print('課外授業を休んでる？: ', name, time, days_of_week[day_of_week], date)
    return time


# noinspection PyUnusedLocal
def update_excel_data(in_file, reference_data, output_file):
    # Read the input Excel file with openpyxl
    output_data = openpyxl.load_workbook(in_file, data_only=False, keep_vba=True)
    input_data = openpyxl.load_workbook(in_file, data_only=True)

    # create a list of children that are 一号 and are in the 課外授業.
    ichigo_kagai_sheet = input_data['1号課外']
    ichigo_kagai = []
    for row in ichigo_kagai_sheet:
        name_val = row[1].value
        if name_val is not None:
            name_val = replace_all_spaces(name_val)
            ichigo_kagai.append(name_val)

    print(ichigo_kagai)

    missing_children = set()
    # Iterate over the input data tabs
    # here I am iterating over the sheet names instead of the worksheet themselves because I will use the sheet names
    # to access the correct file in the difference data.
    for sheet_name in input_data.sheetnames[2:11]:

        # access the sheet we are currently working on
        cur_sheet = input_data[sheet_name]
        out_sheet = output_data[sheet_name]

        # erase any possible spaces in the sheet_name
        new_sheet_name = replace_all_spaces(sheet_name)

        # check to see if tab name exists in reference data
        # if there is no match it is possible the user did not download all the files
        # from hugnote and needs to make sure all classes are selected.
        if new_sheet_name not in reference_data:
            messagebox.showinfo('全てのクラスがダウンロードされてません。', '{}組がダウンロードされてません'.format(new_sheet_name))
            continue
        # Read the reference data for the current tab
        ref_data = reference_data[new_sheet_name]

        # Iterate through the references data
        for i, row in ref_data.iterrows():
            date = row['日付']
            child_name = row['こども氏名']
            child_name = replace_all_spaces(child_name)
            arrive_time = row['出席時刻']
            departure_time = row['帰宅時刻']

            # create day of week num to plug into function to check if the kids are in ichigo_kagai
            clean_date = date.to_pydatetime()
            day_of_week_num = clean_date.weekday()
            # this one is not currently used but may use it in the future to change depending on the day of the week.
            day_of_week_str = clean_date.strftime('%A')

            # remove ':' from time stamp and skip procedure if it is a nan value.
            if isinstance(arrive_time, str):
                arrive_time = arrive_time.replace(':', '')

            if isinstance(departure_time, str):
                departure_time = departure_time.replace(':', '')

            # find the corresponding date(cell row and col) date_coor[0] is the row and date_coor[1] is the column
            date_coor = find_date(cur_sheet, date)

            # check to see if date_coor is empty or is None, if so skip the date. because it can cause errors in the
            # following operations.
            if date_coor is None:
                continue

            # find the corresponding name. This only gives the row because we will use the column numbers from date_coor
            name_coor = find_name(cur_sheet, child_name, date_coor[0])
            # print(i, date_coor, date, new_sheet_name, child_name)
            if not name_coor:
                missing_children.add('{}組の{}'.format(new_sheet_name, child_name))
                print(cur_sheet)
                # print(i, date_coor, date, new_sheet_name, child_name)
                # print(i, name_coor, child_name)

            # check to see if name_coor is an empty list. if so continue to next entry.
            if len(name_coor) == 0:
                continue

            # Write data into cells.
            if isinstance(arrive_time, str):
                arrive_time = int(arrive_time)  # change type to allow for int comparisons
                arrive_time = arr_check_time(arrive_time)  # adjust time for those arriving earlier than 730.
                # Add one to adjust for the 0 index created with enumerate() function
                out_sheet.cell(name_coor[0] + 1, date_coor[1] + 1).value = arrive_time
            if isinstance(departure_time, str):
                departure_time = int(departure_time)  # change type to allow for int comparisons

                ## TODO This process is currently halted due to changes in pricing policy.
                ## TODO will leave here in the case that changes are reversed

                # if ichigo_check(name_coor, cur_sheet):  # check if child is 一号.
                #     departure_time = dep_check_time(departure_time)

                if child_name in ichigo_kagai:  # adjust time if the kids are in 課外授業　and are 一号.
                    departure_time = kagai_ichigo_check_time(child_name, departure_time, day_of_week_num,
                                                             ichigo_kagai_sheet, date)
                # Add one to adjust for the 0 index created with enumerate() function
                out_sheet.cell(name_coor[0] + 1, date_coor[1] + 2).value = departure_time

    if missing_children:
        messagebox.showinfo('以下の子供が見つかりませんでした。',
                            "ハグノートと預かり料金ファイルの子供の漢字が異なってる可能性があります。\n預かり料金ファイルの子供の名前を修正してください。:\n\n" + "\n".join(
                                missing_children))
    else:
        messagebox.showinfo('完了', 'データ転送が完了しました。')
    output_data.save(output_file)
    output_data.close()
    input_data.close()


def recalculate_vba_code(in_file):
    """
    Trigger the calculations in the Excel book externally so that we can access the results in the next step.
    :return:
    """
    app = xw.App(visible=False)
    workbook = xw.Book(in_file)
    workbook.app.calculation = 'automatic'
    workbook.save(in_file)
    workbook.close()
    app.quit()


def find_total_row(sheet: Worksheet) -> list[int]:
    """
    find the rows that have '日計' so that it only iterates through those rows
    """
    ans = []
    for i, row in enumerate(sheet.iter_rows(), start=1):
        cell_value = row[2].value
        if isinstance(cell_value, str):
            cell_value = replace_all_spaces(cell_value)
            if cell_value == '日計':
                ans.append(i)
    return ans


def mark_charges_with_pink(in_file: str) -> None:
    """
    finds cells that have numbers in them which indicates that we have charged the parents money for staying late.
    Then it fills in the cell with a light pink color so it easy to identify where these charges are.
    :param in_file:
    :return None:
    """
    output_data = openpyxl.load_workbook(in_file, data_only=False, keep_vba=True)
    input_data = openpyxl.load_workbook(in_file, data_only=True)
    for in_work_sheet, out_work_sheet in zip(input_data.worksheets[2:11], output_data.worksheets[2:11]):
        cells = []
        # check for individual charges
        for i, row in enumerate(in_work_sheet.iter_rows(min_row=6)):
            for idx, cell in enumerate(row[5::4]):
                if isinstance(cell.value, int) and cell.value >= 100:
                    # this is +6 because workbook objects are 1 indexed but when slicing withe [5::4] it is 0 indexed
                    cells.append((i + 6, (idx * 4) + 6))

        # check for total charges
        rows_to_check = find_total_row(in_work_sheet)
        for i, row in enumerate(in_work_sheet.iter_rows(min_row=4), start=4):
            if i in rows_to_check:
                for idx, cell in enumerate(row[3:]):
                    if isinstance(cell.value, int) and cell.value >= 100:
                        print('total for day: ', cell.value)
                        cells.append((i, idx + 4))

        for cell in cells:
            row = cell[0]
            col = cell[1]
            lavender = 'ffccff'
            out_work_sheet.cell(row=row, column=col).fill = PatternFill(fgColor=lavender, fill_type="solid")
    output_data.save(in_file)
    output_data.close()
    input_data.close()
    messagebox.showinfo('完了', '追加料金があったセルの色塗りが完了しました。')


def import_ref_data(choice: str) -> dict:
    # noinspection GrazieInspection
    """
        Return the reference files all saved into a dictionary that will be imported from a zip file or a regular
        directory depending on the choice of the user.

        :param choice: A string indicating the user's choice. Expected values are 'yes' for zip files
                       and 'no' for regular directories
        :return: A dictionary object containing the contents of all the individual sheets coded to the class name.

        Note:
        -- The function will prompt the user to open a zip file or a folder depending on the choice the user made
           in a previous prompt.
        """
    class_names = ['ひよこ', 'ひつじ', 'うさぎ', 'だいだい',
                   'もも', 'みどり', 'き', 'あお', 'ふじ']
    ref_files = {}
    # import data from reference file. choose method depending on whether user wants to use zip file or not.
    if choice == 'no':
        directory_path = filedialog.askdirectory(title='ダウンロードした打刻表のフォルダを選択してください。')
        files = os.listdir(directory_path)

        # create dictionary to store path names for reference files
        for class_name in class_names:
            for file_name in files:
                file_path = os.path.join(directory_path, file_name)  # create the new file path
                if os.path.isfile(file_path) and class_name in file_path:
                    ref_files[class_name] = pd.read_csv(file_path, parse_dates=['日付'])

    elif choice == 'yes':
        zip_path = filedialog.askopenfilename(title='ダウンロードした打刻表のZIPフォルダを選択してください。',
                                              filetypes=[('Zip Files', '*.zip')])
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_file_names = zip_ref.namelist()
            for class_name in class_names:
                for zip_file_name in zip_file_names:
                    # encode and decode to that we can properly access the file names properly. if not we end up with
                    # odd characters like é╨éµé▒. I happen to find the right encoding, but if the person producing the
                    # zipfile changes the encoding this can cause an error to occur.
                    decoded_name = zip_file_name.encode('cp437').decode('shift_jis')
                    if class_name in decoded_name:
                        unzipped_data = zip_ref.read(zip_file_name).decode('utf-8')
                        ref_files[class_name] = pd.read_csv(StringIO(unzipped_data), parse_dates=['日付'])

            # # Extract zipped files
            # zip_file_names = zip_ref.namelist()
            # unzipped_files = []
            # for file in zip_file_names:
            #     unzipped_files.append(zip_ref.read(file).decode('utf-8'))
            # for class_name in class_names:
            #     for file in unzipped_files:
            #         if class_name in file:
            #             # noinspection PyTypeChecker
            #             # stringIO allows passing in the unzipped file which is already a string instead of saving it as
            #             # a file and then passing in the temporary file address.
            #             ref_files[class_name] = pd.read_csv(StringIO(file), parse_dates=['日付'])
    print(ref_files.keys())
    return ref_files


def range_adjustment(ranges: list[list[int]]) -> list[list[int]]:
    # noinspection GrazieInspection
    """
        Due to some issues with not being able to calculate the values in some cells in the Excel sheet I have created
        this work around function, The list contains two lists of two integers representing ranges. because of the
        mentioned issue only the first range is correct, as well as the beginning of the second range, but not the end
        of the second range.  Using the width of the first range I can then calculate what the end of the second range
        should be.

        :param ranges: A list of two lists with two integers each representing two ranges.
        :return: the adjusted ranges with the second range being corrected.

        Example:
        --------
        >> range_adjustment([[5, 21], [30,30]])
        [[5, 21],[30, 46]]
        """
    first_range = ranges[0]
    range_width = first_range[1] - first_range[0]
    ranges[1][1] = ranges[1][0] + range_width
    return ranges


def find_name_range(sheet: Worksheet) -> list[list[int]]:
    # noinspection GrazieInspection
    """
        locate the rows that need to be searched in order to fill in cells for the absent children.
        :param sheet: The current sheet of an Excel workbook
        :return: a list of tuples indicating the two ranges of rows that needs to be searched for blank cells
                example -> [(6,27), (35, 56)] the second number of the tuple is +1 to account for python ranges
                not being inclusive
        """
    ans = []
    start = False
    temp = []
    for i, row in enumerate(sheet.iter_rows()):
        if row[2].value == '氏名':
            start = True
            # print(1, temp, start)
            temp.append(i + 2)  # its 2 because 0 index plus I want to start at the row after this one
            # print(2, temp, start)

        if start is True and (row[2].value is None or row[2].value == 0):
            start = False
            # print(3, temp, start)
            temp.append(i + 1)
            # print(4, temp, start)
            ans.append(temp)
            temp = []
            # print(5, temp, start)

    # A temporary fix to work around not being able to read the function results in the Excel file.

    return range_adjustment(ans)


def mark_absent(in_file: str) -> None:
    # noinspection GrazieInspection
    """
        Go through the workbook and fill in sections with '休み' where both arrival time and departure time are missing.
        Also mark with yellow where only one of the arrival or departure times is missing. Do nothing to cells that have
        both.

        :param in_file: an Excel workbook
        :return: None since the changes will be happening in place.
        """
    # technically I may not need to have two separate files created because the section that I will be checking is not
    # generated by the vba code, so I should be able to only use the output data (because I will be using this to save
    # so that the vba code stays intact, if I save the input_data the vba code is lost) but just to stay consistent with
    # other functions I will be using both.
    output_data = openpyxl.load_workbook(in_file, data_only=False, keep_vba=True)
    input_data = openpyxl.load_workbook(in_file, data_only=True)
    for in_work_sheet, out_work_sheet in zip(input_data.worksheets[2:11], output_data.worksheets[2:11]):
        search_rows = find_name_range(in_work_sheet)
        for rows in search_rows:
            for row_num in range(rows[0], rows[1]):
                row = in_work_sheet[row_num]
                for idx, cell in enumerate(row[3:59:4]):
                    # adjusting index because enumerate step(skip) doesn't take into account the skipped indices.
                    cell_num = idx * 4 + 3
                    merged = in_work_sheet.merged_cells
                    if cell.coordinate in merged:  # skip any merged cells
                        continue
                    if cell.value is None and row[cell_num + 1].value is None:
                        out_work_sheet[row_num][cell_num].value = '休み'
                        out_work_sheet[row_num][cell_num + 1].value = '休み'

    output_data.save(in_file)
    output_data.close()
    input_data.close()
    messagebox.showinfo('完了', '空欄のセルに休みの書き込みが完了しました。')


# add result to the end of the file name
def new_file_path(path: str, added_text: str = 'result') -> str:
    """
    This function creates a new name for the path of a save file. This is to avoid saving over the original Excel file
    that was used to create the new one. It places a new text between the name and the extension name. If no added_text
    is provided the default 'result' will be used.
    :param path: The path of the original Excel file.
    :param added_text: The text that will be added inbetween the name and the extension name of the original path.
    :return: The newly formed name path where the new Excel file will be saved to.
    """
    idx = -1
    # find the last occurrence of '.' to find where the file type extension is located.
    for i, char in enumerate(path):
        if char == '.':
            idx = i
    if idx == -1:
        return path + added_text
    else:
        return path[:idx] + added_text + path[idx:]


if __name__ == '__main__':

    # create file paths by asking the user.

    # Create the Tkinter root window
    root = tk.Tk()
    root.withdraw()  # Hide the root window

    # prompt user for input file
    input_file = filedialog.askopenfilename(title='①★追加預かり料金原本を選択してください。')

    # TODO import function new_file_path from create billing so that we can easily adjust the new name as well as to
    # TODO to unify how the new file name is created.
    # Generate output file name
    result_file = new_file_path(input_file, '★★作成シート★★')

    # Ask user if they would like to use a zip file or if they already have unzipped the file.
    # and then import the reference files into a dictionary to be used later on.

    # this code is temporarily commented out so that the the program expects to open a zipfile
    # result_choice = messagebox.askquestion('一つを選んでください',
    #                                        'ZIPファイルを使ってデータ転送をしますか？\n(展開がもう済んでいて普通のファイルを開けたい場合は no を選択してください。)',
    #                                        icon='warning')

    # replace result choice with 'yes' so it chooses zip files automatically.
    reference_files = import_ref_data('yes')
    update_excel_data(input_file, reference_files, result_file)
    recalculate_vba_code(result_file)
    mark_charges_with_pink(result_file)
    mark_absent(result_file)
    # final recalculate so that the file doesn't need to be manually opened and saved in order to use the create billing
    # script which would normally crash unless we manually opened the resulting file and then saved it.
    recalculate_vba_code(result_file)
