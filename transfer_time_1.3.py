"""
Version 1.3

First stable version that transfers drop off and pick up times from one excel file to another that then calculates
the appropriate amount of money to charge.

General features
1 -- This version transfers time from one excel file to another.
2 -- There is a window interface to choose the files needed in this script.
3 -- If there are missing files from the refrerence data (data downloaded from hugh note) it will notify you that not all
     files have been downloaded from hugnote, but will transfer data with what ever files are available.
4 -- If children on the hugnote files(reference files) cannot be found in the 預かり料金表 then it will notify you which
     children cannot be found along with the class name.
5 -- Converts the pick up time if a child is in 課外授業 and is 一号.  Children that are 一号 taking the after school class
     are exempt for charges resulting for being picked up late, up to a certain point.
(new)
6 -- created function to replace spaces between names including the japanese space aka IDEOGRAPHIC SPACE character or
     \u3000 in unicode escape character.  This was done to reduce replication to reduce effort when refactoring code.
"""

import pandas as pd
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl import styles
import os



def replace_all_spaces(words: str) -> str:
    words = words.replace('\u3000', '') # \u3000 is the equivalent to the japanese space. normal space -> ' '
    words = words.replace(' ', '')                                                    # japanese space -> '　'
    return words



def find_date(tab: Workbook, date: datetime):
    '''find the row and column (essentially the corordinates) of the matching date.'''
    for i, row in enumerate(tab.iter_rows()):
        for idx, cell in enumerate(row):
            if cell.value == date:
                return [i, idx]
    return None



def find_name(tab1: Workbook, name: str, date_row: int) -> list[int]: # speficically returning one int in the format of a list to avoid out of index errors.
    '''find the row number of where the childs name is located in thr workbook'''
    ans = []
    for i, row in enumerate(tab1.iter_rows()):
        if type(row[2].value) == str:
            cell_name = replace_all_spaces(row[2].value)
            if cell_name == name:
                ans.append(i)
    if date_row < 10:
        return ans[:1] # returning just one so make sure we only place the time correctly for the corresponding date
    else:
        return ans[-1:] # also using [:1] and [-1:] so an error is not raised when the list is empty



def arr_check_time(time: str) -> int:
    '''Convert the arival time so that if anyone arrives before 7:30 it is set to 7:30'''
    time = int(time)
    if  time < 730:
        time = 730
    return time

def dep_check_time(time: str) -> int:
    '''Convert the departure time so that if they are 9 minuets over the time limit it reverts to the time limit
        So that they are not charged'''
    time = int(time)
    if 1131 <= time <= 1139:
        time = 1130
    if 1431 <= time <= 1439:
        time = 1430
    return time


def kagai_ichigo_check_time(name: str, time: int, day_of_week: int, sheet: Workbook) -> int:
    '''Convert departure time for children that are in 課外授業 and are 一号 to 1430 if they have class that day and
       they leave before the pick up time limit.'''
    days_of_week = ["月曜日", "火曜日", "水曜日", "木曜日", "金曜日", "土曜日", "日曜日"]
    for i, row in enumerate(sheet):
        name_val = row[1].value
        if name_val != None:
            name_val = replace_all_spaces(name_val)
        if name == name_val:
            time_limit = sheet[i+1][2 + day_of_week].value
            if time_limit == None:
                #print('修正なし: {}は{}に課外がありません。'.format(name, days_of_week[day_of_week]))
                return time
            elif 1430 < time <= time_limit:
                print('修正あり:　{} {} --> 1430。'.format(name, time))
                time = 1430
                return time
            elif time > time_limit:
                print('修正なし: {}{}が限度の{}を超えてるため。'.format(name, time, time_limit))
                return time
    print('課外授業を休んでる？: ', name, time, days_of_week[day_of_week], date)
    return time




def update_excel_data(input_file, reference_files, output_file):

    # Read the input Excel file with openpyxl
    output_data = openpyxl.load_workbook(input_file, data_only=False, keep_vba=True)
    input_data = openpyxl.load_workbook(input_file, data_only=True)

    # Read the reference CSV files
    reference_data = {}
    for key, val in reference_files.items():
        reference_data[key] = pd.read_csv(val, parse_dates=['日付'])

    # create a list of children that are 一号 and are in the 課外授業.
    ichigo_kagai_sheet = input_data['1号課外']
    ichigo_kagai = []
    for row in ichigo_kagai_sheet:
        name_val = row[1].value
        if name_val != None:
            name_val = replace_all_spaces(name_val)
            ichigo_kagai.append(name_val)

    print(ichigo_kagai)

    missing_children = set()
    # Iterate over the input data tabs
    for sheet_name in input_data.sheetnames[2:11]:

        # access the sheet we are currently working on
        cur_sheet = input_data[sheet_name]
        out_sheet = output_data[sheet_name]

        # erase any possible spaces in the sheetname
        new_sheet_name = replace_all_spaces(sheet_name)

        # check to see if tab name exists in reference data
        # if there is no match it is possible the user did not download all the files
        # from hugnote and needs to make sure all calsses are selected.
        if new_sheet_name not in reference_data:
            messagebox.showinfo('全てのクラスがダウンロードされてません。', '{}組がダウンロードされてません'.format(new_sheet_name))
            continue
        # Read the reference data for the current tab
        ref_data = reference_data[new_sheet_name]

        # Iterate through the refference data
        for i, row in ref_data.iterrows():
            global date
            date = row['日付']
            child_name = row['こども氏名']
            child_name = replace_all_spaces(child_name)
            arrive_time = row['出席時刻']
            departure_time = row['帰宅時刻']

            # create day of week num to plug into function to check if the kids are in ichigo_kagai
            clean_date = date.to_pydatetime()
            day_of_week_num = clean_date.weekday()
            day_of_week_str = clean_date.strftime('%A')

            # remove ':' from time stamp and skip procedure if it is a nan value.
            if isinstance(arrive_time, str):
                arrive_time = arrive_time.replace(':', '')

            if isinstance(departure_time, str):
                departure_time = departure_time.replace(':', '')

            #find the corresponding date(cell row and col) date_coor[0] is the row and date_coor[1] is the column
            date_coor = find_date(cur_sheet, date)

            # check to see if date_coor is empty or is None, if so skip the date. because it can cause errors in the fillowing operations.
            if date_coor == None:
                continue

            # find the corresponding name. This only gives the row because we will use the column numbers from date_coor
            name_coor = find_name(cur_sheet, child_name, date_coor[0])
            if not name_coor:
                missing_children.add('{}組の{}'.format(new_sheet_name, child_name))
                print(i, date_coor, date)
                print(i, name_coor, child_name)

            # check to see if name_coor is an empty list. if so continue to next entry.
            if len(name_coor) == 0:
                continue

            # Write data into cells.
            if isinstance(arrive_time, str) :
                adj_arrive_time = arr_check_time(arrive_time)
                out_sheet.cell(name_coor[0] + 1, date_coor[1] + 1).value = adj_arrive_time # Add one to adjust for the 0 index created with the enumreate() function
            if isinstance(departure_time, str):
                adj_departure_time = dep_check_time(departure_time)
                if child_name in ichigo_kagai: # adjust time if the kids are in 課外授業　and are 一号.
                    adj_departure_time = kagai_ichigo_check_time(child_name, adj_departure_time, day_of_week_num, ichigo_kagai_sheet)
                out_sheet.cell(name_coor[0] + 1, date_coor[1] + 2).value = adj_departure_time # Add one to adjust for the 0 index created with the enumreate() function

    if missing_children:
        messagebox.showinfo('以下の子供が見つかりませんでした。', "ハグノートと預かり料金ファイルの子供の漢字が異なってる可能性があります。\n預かり料金ファイルの子供の名前を修正してください。:\n\n" + "\n".join(missing_children))
    else:
        messagebox.showinfo('完了', 'データ転送が完了しました。')





    output_data.save(output_file)


# create file paths by asking the user.

# Create the Tkinter root window
root = tk.Tk()
root.withdraw()  # Hide the root window


# prompt user for input file
input_file = filedialog.askopenfilename(title = '預かり料金表を選択してください。')
directory_path = filedialog.askdirectory(title = 'ダウンロードした打刻表のフォルダを選択してください。')
files = os.listdir(directory_path)

# Generate output file name
output_file = os.path.splitext(input_file)[0] + "_result.xlsm"

# create dictionary to store path names for reference files
class_names = ['ひよこ', 'ひつじ', 'うさぎ', 'だいだい',
              'もも', 'みどり', 'き', 'あお', 'ふじ']
reference_files = {}
for class_name in class_names:
    for file_name in files:
        file_path = os.path.join(directory_path, file_name) # create the new file path
        if os.path.isfile(file_path) and class_name in file_path:
            reference_files[class_name] = file_path



update_excel_data(input_file, reference_files, output_file)

