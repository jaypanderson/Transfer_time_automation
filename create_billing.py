"""
A script that will take the results from the transfer_time script and insert key data into a billing
form that will be delivered to the parents for extra charges incurred during that month.
"""

from __future__ import annotations
from typing import Union
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from datetime import datetime
import openpyxl
from openpyxl.styles import Color
from openpyxl import Workbook
# from openpyxl import styles
from openpyxl.styles import PatternFill
import xlwings as xw
import zipfile
from io import StringIO
import os
from collections import defaultdict
from transfer_time import replace_all_spaces
from transfer_time import find_name_range
from transfer_time import range_adjustment
from itertools import chain
from copy import copy
from collections import Counter
from itertools import zip_longest
from openpyxl.utils.cell import range_boundaries


from openpyxl.worksheet.worksheet import Worksheet


# TODO open_billing_file and open_excel_file can probably be merged into one function just make it another option.
# open file that will be used create billing docs.
def open_billing_file(option: int) -> str:
    """
    This functions prompts the user open up the base Excel file for the billing form and the tally sheet. Depending on
    the option variable there are different prompts asking for the different files.
    :param option: and integer represented which file should be selected when prompted to open a file.  This option
    then changes the message in the prompt menu so the user knows which file they should be selecting.
    :return: The file path for the file the user has selected.
    """
    if option == 1:
        title = '料金明細票を選択してください。'
    elif option == 2:
        title = '料金集計を選択してください。'
    else:
        title = 'incorrect option chosen'

    return filedialog.askopenfilename(title=title)



# open Excel file that has the information about the extra charges.
def open_excel_file() -> str:
    """
    This funtion prompts the user to open the Excel file that was created with the transfer_time.py script.  It then
    takes finds the file path and returns it.
    :return: The file path for the file the user has selected.
    """
    return filedialog.askopenfilename(title='打刻表を選択してください。')


# add result to the end of the file name
def new_file_path(path: str, added_text: str = 'result') -> str:
    """
    This function creates a new name for the path of a save file. This is to avoid saving over the original Excel file
    that was used to create the new one. It places a new text between the name and the extension name. If no added_text
    is provided the default 'result' will be used.
    :param path: The path of the original Excel file.
    :param added_text: The text that will be added inbetween the name and the extension name of the original path.
    :return: The newly formed name path where the new Excel file will be saved to.
    """
    idx = path.find('.')
    ans = path[:idx] + added_text + path[idx:]
    return ans


# TODO add safeguards so the program doesn't crash when the use chooses the wrong file, but instead re-prompts the user
# TODO to open up the correct one.
# create list or dict with all the extra charges for each child.
def count_charges() -> defaultdict:
    """
    This function looks through an Excel file that was creating using the transfer_time.py script to organize all the
    extra charges into a dictionary.  The function is set up so that it will prompt the user to open up a file. The User
    can potentially choose the wrong file, which will cause the program to crash.
    :return: A nested dictionary that contains all the information needed about the extra charges to create the billing
    documents.
    """
    file_path = open_excel_file()
    book = openpyxl.load_workbook(file_path, keep_vba=False, data_only=True)
    charges = defaultdict(lambda : defaultdict(list))
    # iterate through the sheets
    for sheet_name in book.sheetnames[2:11]:
        sheet_name = replace_all_spaces(sheet_name) # replace spaces
        sheet = book[sheet_name]
        row_ranges = find_name_range(sheet)
        # iterate through the two ranges (its like doing two chained iterations. but this way it's easier to calculate
        # the date row)
        for row_range in row_ranges:
            start = row_range[0]
            end = row_range[1]
            for idx, row in enumerate(sheet.iter_rows(min_row=start, max_row=end-1)):
                name = row[2].value
                for i, cell in enumerate(row[5::4]):
                    price = cell.value
                    # TODO refactor so that the code is more organized.
                    try:
                        price = int(price)
                    except TypeError:
                        continue
                    except ValueError:
                        # TODO add proper warnings to that we can identify where there was an error with the formular.
                        print('数式に間違いがあるかもしれません', price, name, i*4 + 6)
                        continue

                    if price is not None and price >= 100:
                        date_row = start - 4
                        date_col = i*4 + 4 # (its 6 for the price but 2 less for the column that has the date.)
                        date = sheet.cell(row=date_row, column=date_col).value
                        date = str(date)[0:10]
                        arriv_row = start + idx
                        arriv_col = i * 4 + 4
                        dept_row = start + idx
                        dept_col = i * 4 + 5
                        arrival = sheet.cell(row=arriv_row, column=arriv_col).value
                        departure = sheet.cell(row=dept_row, column=dept_col).value
                        charges[sheet_name][name].append((price, arrival, departure, date))
    print(charges)
    return charges


# Set the color of the tab for a sheet.
def set_color(sheet: Worksheet, class_name: str) -> None:
    """
    This function sets the tab color of the newly created sheet to a predetermined color based on the class.
    :param sheet: The newly created sheet
    :param class_name: The name of the class this work sheet will be working with to creating the billing form of a
    particular child.
    :return: None
    """
    color_map = {'ひよこ': 'FFFFFCE6', 'ひつじ': 'FFE6FFE6', 'うさぎ': 'FFE6FFFF', 'もも': 'FFFFC0CB',
                 'だいだい': 'FFFFA500', 'き': 'FFFFFF00', 'みどり': 'FF008000', 'あお': 'FF0000FF', 'ふじ': 'FF800080'}
    color = Color(rgb=color_map[class_name])
    sheet.sheet_properties.tabColor = color


# copy the contents of one sheet to another.
# noinspection PyDunderSlots,PyUnresolvedReferences
def copy_sheet(sheet: Worksheet, new_sheet: Worksheet) -> None:
    """
    copy the contents from one sheet to another sheet. This function only copies the contents of the cells and its
    style, other aspects of the sheet are copied with other functions.
    :param sheet:The base sheet from which the cells are copied.
    :param new_sheet: The new sheet where the cell contents and style are pasted.
    :return: None
    """
    for row in sheet:
        for cell in row:
            new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
                new_cell.comment = copy(cell.comment)


# replicate the cell merges from base template.
def merge_cells(sheet: Worksheet, new_sheet: Worksheet) -> None:
    """
    When a new sheet is created and everything is copied, the one thing that will not be copied is the location of where
    there are merged cells.  This function iterates over all the merged cells of the base worksheet and then merges
    those cells in the new sheet.
    :param sheet: The base work sheet this function will use to iterate over the location of the merged cells.
    :param new_sheet: The new work sheet where the function will merge the cells from the base work sheet.
    :return: None
    """
    for merged_cell_range in sheet.merged_cells.ranges:
        new_sheet.merge_cells(str(merged_cell_range))


# copy the part of the worksheet that will be printed onto the new sheet.
def copy_print_area(sheet: Worksheet, new_sheet: Worksheet) -> None:
    """
    A certain area of a page is selected as the default area to be printed when the print button is pressed. To avoid
    having to change the print area for every single new sheet this function copies the print area attribute from the
    base sheet to the new sheet.
    :param sheet: the base sheet from which we will copy the print area attribute.
    :param new_sheet: The new sheet where we will paste the print area.
    :return: None
    """
    if sheet.print_area:
        new_sheet.print_area = sheet.print_area

# copy the width and height of cells
def copy_dimensions(sheet: Worksheet, new_sheet: Worksheet) -> None:
    """
    A function that changes the new sheet's size of the cells to match that of a another sheet. Currently, it is set up
    so that it only copies the width of the cells.  If the height also needs to be copied, then Uncomment that portion
    of the code and vice versa.
    :param sheet: The sheet from which the cell dimensions will be copied from
    :param new_sheet: The sheet where the dimensions will be pasted into.
    :return: None
    """
    for row, col in zip_longest(sheet.row_dimensions, sheet.column_dimensions):
        # if row is not None:
        #     new_sheet.row_dimensions[row].height = sheet.row_dimensions[row].height
        if col is not None:
            new_sheet.column_dimensions[col].width = sheet.column_dimensions[col].width



# function to insert the child's information into the billing document
def insert_name_date(sheet: Worksheet, year: int, month: int, class_name: str, child_name: str) -> None:
    """
    In the base template worksheet there are characters that demark certain information in the document. In this
    instance the following are used ['%', '#', '?', '@', '&', '$'], each representing a certain data.
    :param sheet: The new worksheet that was generated to insert this data.
    :param year: The current year, where it is represented as '%'
    :param month: The current month, where it is represented as '#'
    :param class_name: The name of the class for the child with extra charges. The class name is represented as '@'
    while the age is represented as '?' the age is calculated by using the class age map.
    :param child_name: The name of the child with extra charges. The child's name is represnted as the first name, then
    a japanese space(/u3000) and then the first name. ex: '田中　太郎'.  The last name is represented as '&' and the first
    name is represented as '$'.
    :return: None
    """
    class_age_map = {'あお': '5', 'ふじ': '5', 'き': '4', 'みどり': '4', 'だいだい': '3', 'もも': '3',
                     'うさぎ': '2', 'ひつじ': '1', 'ひよこ': '0'}
    full_name = child_name.split('　')
    last = full_name[0]
    first = full_name[1]
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if '%' in cell.value:
                cell.value = cell.value.replace('%', str(convert_reiwa(year, month)))
            if '#' in cell.value:
                cell.value = cell.value.replace('#', str(month))
            if '?' in cell.value:
                cell.value = cell.value.replace('?', class_age_map[class_name])
            if '@' in cell.value:
                cell.value = cell.value.replace('@', class_name)
            if '&' in cell.value:
                cell.value = cell.value.replace('&', last)
            if '$' in cell.value:
                cell.value = cell.value.replace('$', first)


# helper function to find the key that has the highest value associated with it.
def find_max(counts: Counter) -> int:
    """
    This function is intended to return the key that has the highest value associated with it in a Counter dictionary
    object.
    :param counts: Counter object containing the counts of each key. In this script this function finds the most
    frequent year and month.
    :return: the year or month that has the highest frequency.
    """
    highest = 0
    ans = None
    for i in counts:
        if counts[i] > highest:
            highest = counts[i]
            ans = i
    return ans


# function to find the most common year and month.
def find_year(charges: dict) -> tuple[int, int]:
    """
    The idea behind this function is to get the current year and month with allowing for some mistakes in input.
    Instead of checking for any single cell to get the year which can lead to a mistake where one cell has a certain
    year and the other has another, this function finds the most common year and month.
    :param charges: The dictionary that contains all the information about the extra charges for all the children.
    :return:  tuple with two integers, where the first represents the year, and the second re@presents the month.
    [year, month] or [2023, 10]
    """
    years = []
    months = []
    for i in charges:
        for j in charges[i]:
            for data in charges[i][j]:
                year = data[3].split('-')[0]
                years.append(year)
                month = data[3].split('-')[1]
                months.append(month)

    year_ans = find_max(Counter(years))
    month_ans = find_max(Counter(months))
    return int(year_ans), int(month_ans)


# Convert from Gregorian (Western) calendar years into Japanese calendar years.
def convert_reiwa(year: int, month: int) -> str:
    """
    Convert from Gregorian (Western) calendar years into Japanese calendar years.  The Reiwa year system started
    at 2018　and the first year is represented as '元' instead of the number 1.
    :param year: The current year
    :param month: The current month
    :return: a string of the year converted into the japanese calendar year.
    """
    reiwa = year - 2018
    if month in [1, 2, 3]:
        reiwa -= 1
    if reiwa == 1:
        reiwa = '元'

    return str(reiwa)


# function to copy the contents of a row into another.
def copy_row_style(sheet: Worksheet, row_num: int, new_row_num: int) -> None:
    """
    This function copies the style of a row into another.  It is meant to copy and replicate the row as
    much as possible.  Not all the attributes are copied over, just the main ones.  This function is indented to be
    used when a new row is inserted, so that the new row looks the same as the previous row.
    :param sheet: Worksheet that we are working on
    :param row_num: the row number from which the function will copy the style.
    :param new_row_num: The row number of where the function will paste the style.
    :return: None
    """
    for row, new_row in zip(sheet.iter_rows(min_row=row_num, max_row=row_num),
                            sheet.iter_rows(min_row=new_row_num, max_row=new_row_num)):
        for cell, new_cell in zip(row, new_row):
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)


# convert the dates into a preferred format.
def convert_date(date: str) -> str:
    """
    Change the format of the date to the desired format. In this particular case we want to convert the date format from
    '2023-10-07' to '10月7日'. Notice how the 0 is gone from the day. this must also be the case for the month
    so if we had '2023-04-03' it should be converted to '4月3日'
    :param date: a date string in the format of year dash month dash day i.e. '2023-10-07'
    :return: the desired format of a date string i.e. '10月7日'.
    """
    ye_mo_da = date.split('-')
    month = ye_mo_da[1]
    date = ye_mo_da[2]

    # I'm converting to int and then inserting into f string because the leading 0 in something like
    # '10月07日' was not desirable. instead we want something like '10月7日'. So converting to int gets
    # rid of the leading 0.
    return f'{int(month)}月{int(date)}日'


# change the time to the desired format.
def format_time(time: int) -> str:
    """
    The arrival time and departure time are formatted aas 1034 or 1645 in the worksheet that keeps track of the times
    of the individual kids.  This function converts it into the format where it has ':' in the middle such as
    10:34 or 16:45. Also returns 'Time' error if the length of time is not 3 or 4. Because this would indicate that
    the time was inserted incorrectly.
    :param time: a string of the time in the format of 1645
    :return: a string of the time in the format of 16:45
    """
    chars = list(str(time))
    if 3 > len(chars) > 4:
        return 'Time error'
    chars.insert(-2, ':')
    return ''.join(chars)


# function to insert data into the billing form
def insert_data(sheet: Worksheet, row: int, month: int, price: int, arrival: int, departure: int, date: str) -> None:
    """
    This function inserts the data into the billing documents at their respective rows.  This function is called from
    the create_billing function and it not to be confused with insert_tally_data. This function should be called for
    every day the child has an extra charge.
    :param sheet: The sheet that was created to insert the data
    :param row: The row in which we cant to insert the data into
    :param month: The month of the current billing cycle
    :param price: The accumulated charge for that particular day.
    :param arrival: The time that the child arrived at the school.
    :param departure: The time that the child departed the school.
    :param date: The date at with the extra charges incurred.
    :return: None
    """
    for cells in sheet.iter_rows(min_row=row, max_row=row):
        cells[1].value = f'{month}月分預かり保育料金'
        cells[3].value = convert_date(date)
        cells[4].value = format_time(arrival)
        cells[5].value = format_time(departure)
        cells[6].value = price


# merge specified cells
def merge_specific_cells(sheet: Worksheet, new_row_num: int, start_col: str, end_col: str) -> None:
    """
    merge specified cells. This particular cell merge function only merges columns in a single cell due to the nature
    of merging needed in this document.
    :param sheet: The Worksheet that is currently being worked on.
    :param new_row_num: The row that will be used as the start_row and end_row to merge cells. For this function they
    are the same thing.
    :param start_col: The upper range for which columns will be merged. use a string such as 'D' or 'G'.
    :param end_col: The lower range for which columns will be merged.  use a string such as 'D' or 'G'.
    :return: None
    """
    merge_range = f'{start_col}{new_row_num}:{end_col}{new_row_num}'
    sheet.merge_cells(merge_range)


def adjust_merged_cells(sheet: Worksheet, new_row: int) -> None:
    """
    Because openpyxl has no build in way of adjusting the merged cells when a row is inserted this function was created.
    It works by taking in the location of where the row was inserted and then adjusting any merged cells bellow that row.
    That is this function is build to incrementally shift the merged cells every time it is called.
    :param sheet: The worksheet that we need to adjust the merged cells.
    :param new_row: the location at which a new row was inserted.
    :return: None
    """
    new_merged_ranges = []
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))

        if min_row > new_row + 1:
            min_row += 1
            max_row += 1
        # TODO maybe I dont need to get the column letter maybe i can just use the min col it self and max col it self?
        new_range = f'{openpyxl.utils.get_column_letter(min_col)}{min_row}:{openpyxl.utils.get_column_letter(max_col)}{max_row}'
        new_merged_ranges.append(new_range)

    # the code make sure the there are no merged cells so we can reset them it sets the range to an empty list
    sheet.merged_cells.ranges = []

    for new_range in new_merged_ranges:
        sheet.merge_cells(new_range)


# find the location of the number that needs to be recalculated and add the number of rows based on
# how many rows were inserted.
def recalc_number(formula: str, num_rows_inserted: int, cell_range: bool) -> tuple[int, int, int]:
    """
    This function finds the specific number in the formula that need to be changed. Depending on whether the formula
    is a simple cell number or a formula that uses a range of cells it looks for the desired number to change as well
    as its location.

    :param formula: this is the formula that was found in the cell that need to be adjusted.
    :param num_rows_inserted: This basically is the number of days the child was charged extra.  This is because
    for every day the child was charged, there will be a new row inserted to record the charge.
    :param cell_range: a bool argument that determines how we search for the number that need to be change. If True it
    searches the formula assuming that it uses as range of cells.(ex: sum(B23:B25))  If False it searched the formula assuming that it
    uses a single cell. (ex: =B23)
    :return: returns a tuple containing [new_num, start, end]. new_num is the adjusted number after recalculating with
    the numbers of rows inserted, start is the index of where the number started and end is the index for where the
    number ended. these will be used to splice together the new formula that will be inserted into as cell in the
    new sheet.
    """
    start = None
    end = None
    if cell_range is True:
        for i, char in enumerate(formula):
            if char == ':':
                start = i + 2  # it's two, to account for ':' and the column letter.
            if char == ')':
                end = i
    else:
        for i, char in enumerate(formula):
            if char.isalpha():
                start = i + 1
        end = len(formula)

    num = formula[start:end]
    new_num = int(num) + num_rows_inserted
    return new_num, start, end


# apply the new values to the formulas based on how many rows were inserted.
def adjust_formulas(sheet: Worksheet, cells_to_be_adjusted: tuple[tuple[int, int, bool],...], num_rows_inserted: int) -> None:
    """
    Because rows are being inserted, the range of the formulas that tally the total amounts need to adjust for that.
    Normally if we do this in Excel is automatically adjusts it. However, with openpyxl when a row is inserted the
    formulas are not adjusted automatically.  That's why, depending on the number of rows inserted we need to adjust
    the formulas present in the Excel book.

    :param sheet: new sheet that was created for each child that has extra charges.
    :param cells_to_be_adjusted: tuple of tuples containing the rows, columns, and type of formula of the cells that
    need their formulas to be adjusted. ex:((3, 2, True), (10, 5, False)) Each tuple in the tuple represents a cell.
    The first number in the tuple is the row, the second number is the column of the cell that need to be changed and
    the last bool represents if the formula uses a range of cells or just a single cell. (For True it is a
    range =SUM(D3:D10), for False it is a single cell =D43.)
    :param num_rows_inserted: This basically is the number of days the child was charged extra.  This is because
    for every day the child was charged, there will be a new row inserted to record the charge.
    :return:
    """
    for row, column, cell_range in cells_to_be_adjusted:
        formula = sheet.cell(row, column).value

        new_num, start, end = recalc_number(formula, num_rows_inserted, cell_range)
        sheet.cell(row, column).value = formula[:start] + str(new_num) + formula[end:]


# one of three documents that this automation creates.  This one creates the billing documents that will be
# given to the parents with all the individual charges organized by day.
def create_billing_sheets(charges: defaultdict) -> None:
    """
    This function creates the first document that is a billing form that will be given to the parents. It includes
    the individual charges for each day there was a charge.  It also includes formulas that calculate the subtotal of
    the month for an individual.
    :param charges: A dictionary that organizes the charges in a nested format.
    ex:1 {'class_name': {'kid_name': [(charge, arrival_time, departure_time, date), ...]}, ...}
    ex:2 {'だいだい': {'田中　太郎': [(500, 832, 1700, 2023-10-03), ...]}, ...} The names includes the Japanese space
    character than can also be represented as \u3000
    :return: None
    """
    file_path = open_billing_file(1)
    book = openpyxl.load_workbook(file_path, keep_vba=False)
    sheet = book[book.sheetnames[0]]

    year = find_year(charges)[0]
    month = find_year(charges)[1]

    for class_name in charges:
        for kid_name in charges[class_name]:
            print(month, class_name, replace_all_spaces(kid_name))
            new_sheet_name = f'{month}月{class_name}{replace_all_spaces(kid_name)}'
            new_sheet = book.create_sheet(new_sheet_name)
            set_color(new_sheet, class_name)
            copy_sheet(sheet, new_sheet)
            merge_cells(sheet, new_sheet)
            copy_print_area(sheet, new_sheet)
            copy_dimensions(sheet, new_sheet)
            insert_name_date(new_sheet, year, month, class_name, kid_name)

            rows_inserted = 0
            for i, data in enumerate(charges[class_name][kid_name]):
                row_num = 14
                first_insertion_location = 15
                new_row_num = 14 + i
                if i != 0:
                    new_sheet.insert_rows(row_num + 1 + i)
                    rows_inserted += 1
                    merge_specific_cells(new_sheet, row_num + i, 'B', 'C')
                    adjust_merged_cells(new_sheet, new_row_num)
                copy_row_style(new_sheet, row_num, row_num + i)
                insert_data(new_sheet, new_row_num, month, data[0], data[1], data[2], data[3])

            if rows_inserted > 0:
                cells_to_be_adjusted =((16 + rows_inserted, 7, True), (30 + rows_inserted, 4, False))
                adjust_formulas(new_sheet, cells_to_be_adjusted, rows_inserted)
    book.save(new_file_path(file_path, '★★作成シート★★'))


# Calculate the total charge for a single child.
def price_per_child_total(child_charges: defaultdict) -> int:
    """
    A function to calculate the subtotal of extra charges for an individual child incurred during the month. This
    should be called for every child that has extra charges during the month.
    :param child_charges: The value of the innermost part of the charges' dictionary. it is basically the data that is
    needed for all operations in the form of a list of tuples ex1: [(charge, arrival_time, departure_time, date), ...]
    ex2: [(500, 859, 1631, '2023-10-24'), ...]
    :return: The total charges for a given child
    """
    total = 0
    for data in child_charges:
        total += data[0]
    return total

# function that inserts the data into the correct cell.
def insert_tally_data(new_sheet: Worksheet, row: int, class_name: str, kid_name: str, price: int) -> None:
    """
    A function that inserts the data in to their respective cells. This function should be called for every child that
    has extra charges.
    :param new_sheet: The sheet that was created to insert the data
    :param row: The row in which we cant to insert the data into
    :param class_name: The name of the class. ex: 'うさぎ', 'だいだい'
    :param kid_name: The name of the child. ex: '田中　太郎' the names should have a Japanese space between the first
    and last name. The Japanese name can be represented as \u3000
    :param price: The total charges for a child for that particular month. This value is calculated in another function.
    :return: None
    """
    class_age_map = {'あお': '5', 'ふじ': '5', 'き': '4', 'みどり': '4', 'だいだい': '3', 'もも': '3',
                     'うさぎ': '2', 'ひつじ': '1', 'ひよこ': '0'}
    for cells in new_sheet.iter_rows(min_row=row, max_row=row):
        cells[0].value = f'{class_age_map[class_name]}歳児'
        cells[1].value = class_name
        cells[2].value = kid_name
        cells[3].value = price


# Function to add formula that calculates total for each class.
def insert_formula_class_total(new_sheet: Worksheet, rows_inserted: int, first_row: int) -> int:
    """
    A function to create and insert a new formula into a cell to calculate the total charges for each class.
    it is meant to be called after the information for last child from a given class is inserted into the worksheet.
    This function also returns an integer that represents the beginning row of where the child from the next class will
    be inserted so that the total can be calculated.
    :param new_sheet: The new sheet that was created to insert the charges for each child.
    :param rows_inserted: The number of rows that were inserted for a particular class. Basically the number of children
    from a class that have extra charges. This is then added to 3, which is the first row that any child's data was
    inserted, to calculate the bottom range for the formula.
    :param first_row: The first row from which we start counting to calculate the total for each class. This is the
    upper range used in the formular
    :return: Add one to the last_row, so we know where the next range starts.  This will become the first_row when this
    function is called again.
    """
    last_row = 3 + rows_inserted
    new_sheet.cell(row=last_row, column=8).value = f'=SUM(D{first_row}:D{last_row})'
    return last_row + 1


# function to create the second document for billing purposes.
def create_tally_sheet(charges: defaultdict) -> None:
    """
    This function creates the second document that is a tally of all the charges for each child organized into one
    worksheet.  It displays the age, class name, child name, and the total charges for each child. It also calculates
    the total for each class as well as the subtotal of all the classes.
    :param charges: A dictionary that organizes the charges in a nested format.
    ex:1 {'class_name': {'kid_name': [(charge, arrival_time, departure_time, date), ...]}, ...}
    ex:2 {'だいだい': {'田中　太郎': [(500, 832, 1700, 2023-10-03), ...]}, ...} The names includes the Japanese space
    character than can also be represented as \u3000
    :return: None
    """
    file_path = open_billing_file(2)
    book = openpyxl.load_workbook(file_path)

    year = find_year(charges)[0]
    month = find_year(charges)[1]
    sheet = book['base']
    new_sheet_name = f'{year}.{month}'
    new_sheet = book.create_sheet(new_sheet_name)
    copy_sheet(sheet, new_sheet)
    new_sheet.cell(row=2, column=1).value = f'{year}.{month}月'
    merge_cells(sheet, new_sheet)
    copy_dimensions(sheet, new_sheet)

    first = True
    count = 0
    rows_inserted = 0
    first_row = 3 # row number that will be used to set the range for the sum of each class.
    for class_name in charges:
        for kid_name in charges[class_name]:
            print(class_name, kid_name)
            price = price_per_child_total(charges[class_name][kid_name])
            if first is not True:
                new_sheet.insert_rows(count + 3)
                rows_inserted += 1
            copy_row_style(new_sheet, 3, count + 3)
            insert_tally_data(new_sheet, count + 3, class_name, kid_name, price)
            first = False
            count += 1
        # this functions returns an object as well as mutates the new_sheet that is passed in.
        first_row = insert_formula_class_total(new_sheet, rows_inserted, first_row)

    # make adjustments to the coordinates of the cells that need to be adjusted because we are changing it after the
    # rows are inserted.  I could define the rows being inserted before the loop instead of counting during the loop,
    # but I need the intermediate counts of rows_inserted to add formulas along the way that count how much extra
    # charges each class has.
    cells_to_be_adjusted = ((4 + rows_inserted, 4, True), (4 + rows_inserted, 7, True), (4 + rows_inserted, 8, True))
    adjust_formulas(new_sheet, cells_to_be_adjusted, rows_inserted)

    book.save(new_file_path(file_path, '★★作成シート★★'))


# main function to run all the processes I need.  Currently, this only creates two files. The final file still needs
# some thought put into it on whether it should be created by hand or not.
def main():
    charges = count_charges()
    create_billing_sheets(charges)
    create_tally_sheet(charges)


if __name__ == '__main__':
    #count_charges()
    main()
    #testtest(count_charges())
    #find_year(count_charges())
    #convert_reiwa(2024, 4)